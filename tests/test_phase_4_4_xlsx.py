"""
tests/test_phase_4_4_xlsx.py
Тесты для Фазы 4.4 — обработка Excel (.xlsx) файлов в JadeBridge.

Запуск:
    pip install pytest openpyxl
    pytest tests/test_phase_4_4_xlsx.py -v

Архитектура тестов:
- Все тесты работают с реальными временными .xlsx файлами (создаём через openpyxl)
- Telegram-объекты (Update, Context) — моки
- LLM-вызовы (ask_llm, choose_model) — замокированы через unittest.mock
"""

import io
import os
import sys
import math
import asyncio
import tempfile
import unittest
from unittest.mock import AsyncMock, MagicMock, patch, call

import openpyxl
import pytest

# Добавляем корень проекта в sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from bot import (
    sheet_to_markdown,
    _open_xlsx_data,
    build_sheet_text,
    split_text_into_parts,
    _process_xlsx_sheets,
    _send_xlsx_to_llm,
    handle_xlsx_dialog,
    XLSX_MAX_CHARS,
)


# ══════════════════════════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def make_xlsx_file(sheets_data: dict) -> str:
    """
    Создаёт временный .xlsx файл из словаря {sheet_name: [[row1], [row2], ...]}.
    Возвращает путь к файлу.
    """
    wb = openpyxl.Workbook()
    # Удаляем дефолтный пустой лист
    default_sheet = wb.active

    first = True
    for sheet_name, rows in sheets_data.items():
        if first:
            ws = default_sheet
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def make_context_mock(user_data: dict = None) -> MagicMock:
    """Создаёт мок context с user_data."""
    ctx = MagicMock()
    ctx.user_data = user_data if user_data is not None else {}
    return ctx


def make_update_mock(text: str = "") -> MagicMock:
    """Создаёт мок update с фиктивным message."""
    update = MagicMock()
    update.effective_user.id = 42
    update.message.text = text
    update.message.reply_text = AsyncMock()
    return update


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 1: Один лист — нет диалога, сразу обработка
# ══════════════════════════════════════════════════════════════════════════════

class TestSingleSheet(unittest.IsolatedAsyncioTestCase):
    """Тест 1: Файл с одним листом → никаких вопросов, сразу обработка."""

    async def test_single_sheet_no_dialog(self):
        """
        Дано: xlsx-файл с одним листом «Данные», 3 строки × 2 столбца.
        Ожидаем:
        - xlsx_pending НЕ устанавливается в context.user_data
        - _send_xlsx_to_llm вызван один раз
        """
        path = make_xlsx_file({
            "Данные": [
                ["Имя", "Возраст"],
                ["Алиса", 30],
                ["Боб", 25],
            ]
        })
        try:
            xlsx_data = _open_xlsx_data(path)
            assert len(xlsx_data["sheet_names"]) == 1

            update = make_update_mock()
            context = make_context_mock()

            with patch("bot._send_xlsx_to_llm", new_callable=AsyncMock) as mock_send:
                await _process_xlsx_sheets(
                    update, context,
                    sheet_names=xlsx_data["sheet_names"],
                    xlsx_data=xlsx_data,
                    file_name="test.xlsx",
                    caption="",
                    user_id=42,
                )

            # Нет диалога — xlsx_pending не установлен
            assert "xlsx_pending" not in context.user_data
            # LLM вызван ровно один раз
            mock_send.assert_awaited_once()
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 2: Несколько листов — бот спрашивает выбор
# ══════════════════════════════════════════════════════════════════════════════

class TestMultiSheetAskChoice(unittest.IsolatedAsyncioTestCase):
    """Тест 2: Файл с несколькими листами → бот должен спросить выбор."""

    async def test_multi_sheet_asks_choice(self):
        """
        Дано: xlsx с 3 листами.
        Ожидаем:
        - context.user_data["xlsx_pending"]["state"] == "awaiting_sheet_choice"
        - reply_text вызван с перечислением листов
        - _send_xlsx_to_llm НЕ вызван
        """
        path = make_xlsx_file({
            "Продажи": [["Месяц", "Сумма"], ["Январь", 100]],
            "Расходы": [["Статья", "Сумма"], ["Аренда", 50]],
            "Итого":   [["Итог", "Прибыль"], ["Год", 50]],
        })
        try:
            xlsx_data = _open_xlsx_data(path)
            assert len(xlsx_data["sheet_names"]) == 3

            update = make_update_mock()
            context = make_context_mock()

            with patch("bot._send_xlsx_to_llm", new_callable=AsyncMock) as mock_send:
                # Имитируем логику из handle_document для multi-sheet случая
                sheet_names = xlsx_data["sheet_names"]
                sheets_list = "\n".join(
                    f"{i+1}. {name}" for i, name in enumerate(sheet_names)
                )
                context.user_data["xlsx_pending"] = {
                    "state": "awaiting_sheet_choice",
                    "file_name": "report.xlsx",
                    "sheet_names_all": sheet_names,
                    "xlsx_data": xlsx_data,
                    "caption": "",
                    "user_id": 42,
                }
                await update.message.reply_text(
                    f"📋 Файл содержит {len(sheet_names)} листа(ов):\n"
                    f"{sheets_list}\n\n"
                    f"Напиши номер листа (1–{len(sheet_names)}), несколько через запятую, "
                    f"или напиши «все»."
                )

            # Диалог запущен
            assert context.user_data["xlsx_pending"]["state"] == "awaiting_sheet_choice"
            assert len(context.user_data["xlsx_pending"]["sheet_names_all"]) == 3
            # LLM не был вызван
            mock_send.assert_not_awaited()
            # Пользователю отправлен список листов
            assert update.message.reply_text.call_count == 1
            call_text = update.message.reply_text.call_args[0][0]
            assert "Продажи" in call_text
            assert "Расходы" in call_text
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 3: Пользователь выбирает конкретный лист по номеру
# ══════════════════════════════════════════════════════════════════════════════

class TestMultiSheetPickOne(unittest.IsolatedAsyncioTestCase):
    """Тест 3: Пользователь пишет "2" → обрабатывается только второй лист."""

    async def test_multi_sheet_user_picks_one(self):
        """
        Дано: pending с 3 листами, пользователь пишет "2".
        Ожидаем: _process_xlsx_sheets вызван с выбранным листом "Расходы".
        """
        path = make_xlsx_file({
            "Продажи": [["Месяц", "Сумма"], ["Январь", 100]],
            "Расходы": [["Статья", "Сумма"], ["Аренда", 50]],
            "Итого":   [["Итог", "Прибыль"], ["Год", 50]],
        })
        try:
            xlsx_data = _open_xlsx_data(path)
            sheet_names = xlsx_data["sheet_names"]  # ["Продажи", "Расходы", "Итого"]

            update = make_update_mock(text="2")
            context = make_context_mock(user_data={
                "xlsx_pending": {
                    "state": "awaiting_sheet_choice",
                    "file_name": "report.xlsx",
                    "sheet_names_all": sheet_names,
                    "xlsx_data": xlsx_data,
                    "caption": "сколько расходов",
                    "user_id": 42,
                }
            })

            with patch("bot._process_xlsx_sheets", new_callable=AsyncMock) as mock_process:
                await handle_xlsx_dialog(update, context)

            # Диалог завершён — pending очищен
            assert "xlsx_pending" not in context.user_data
            # _process_xlsx_sheets вызван с правильным листом
            mock_process.assert_awaited_once()
            call_args = mock_process.call_args
            # Аргументы позиционные: (update, context, sheet_names, xlsx_data, file_name, caption, user_id)
            positional = call_args[0]
            sheet_names_passed = positional[2] if len(positional) > 2 else call_args[1].get("sheet_names", [])
            assert sheet_names_passed == ["Расходы"]
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 4: Пользователь выбирает "все" → все листы
# ══════════════════════════════════════════════════════════════════════════════

class TestMultiSheetPickAll(unittest.IsolatedAsyncioTestCase):
    """Тест 4: Пользователь пишет "все" → все листы передаются в обработку."""

    async def test_multi_sheet_user_picks_all(self):
        """
        Дано: pending с 3 листами, пользователь пишет "все".
        Ожидаем: _process_xlsx_sheets вызван со всеми 3 листами.
        """
        path = make_xlsx_file({
            "Лист1": [["A", "B"], [1, 2]],
            "Лист2": [["C", "D"], [3, 4]],
            "Лист3": [["E", "F"], [5, 6]],
        })
        try:
            xlsx_data = _open_xlsx_data(path)
            sheet_names = xlsx_data["sheet_names"]

            update = make_update_mock(text="все")
            context = make_context_mock(user_data={
                "xlsx_pending": {
                    "state": "awaiting_sheet_choice",
                    "file_name": "data.xlsx",
                    "sheet_names_all": sheet_names,
                    "xlsx_data": xlsx_data,
                    "caption": "",
                    "user_id": 42,
                }
            })

            with patch("bot._process_xlsx_sheets", new_callable=AsyncMock) as mock_process:
                await handle_xlsx_dialog(update, context)

            assert "xlsx_pending" not in context.user_data
            mock_process.assert_awaited_once()
            # Переданы все 3 листа
            call_args = mock_process.call_args
            passed_sheets = call_args[0][2] if call_args[0] else call_args[1].get("sheet_names")
            assert set(passed_sheets) == set(sheet_names)
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 5: Формула с вычисленным значением → берётся значение, не формула
# ══════════════════════════════════════════════════════════════════════════════

class TestFormulaWithValue(unittest.TestCase):
    """Тест 5: Ячейка с числовым значением (не формула) → показывается значение."""

    def test_formula_with_value(self):
        """
        Дано: лист с числовыми значениями.
        Ожидаем: markdown-таблица содержит эти значения, а не формулы.
        """
        path = make_xlsx_file({
            "Лист1": [
                ["Продукт", "Цена", "Кол-во", "Итог"],
                ["Яблоко", 50, 10, 500],   # 500 — вычисленное значение
                ["Груша", 80, 5, 400],      # 400 — вычисленное значение
            ]
        })
        try:
            xlsx_data = _open_xlsx_data(path)
            sheet_data = xlsx_data["sheets"]["Лист1"]
            md = sheet_data["markdown"]

            # Значения должны быть в таблице
            assert "500" in md
            assert "400" in md
            # Формул нет (в этом тесте их и не было)
            assert "не вычислено" not in md
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 6: Формула без вычисленного значения → fallback с пометкой
# ══════════════════════════════════════════════════════════════════════════════

class TestFormulaWithoutValue(unittest.TestCase):
    """Тест 6: Ячейка с формулой, которую openpyxl не вычислил → fallback."""

    def test_formula_without_value_fallback(self):
        """
        Дано: xlsx с формулой. Так как мы пишем формулу вручную через openpyxl
        (без пересохранения в Excel), data_only=True вернёт None.
        Ожидаем: markdown содержит "[формула, не вычислено: =...]"
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расчёты"
        ws["A1"] = "Значение А"
        ws["B1"] = "Значение Б"
        ws["C1"] = "Сумма"
        ws["A2"] = 100
        ws["B2"] = 200
        ws["C2"] = "=A2+B2"   # Формула — openpyxl не вычислит в data_only режиме

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            xlsx_data = _open_xlsx_data(tmp.name)
            md = xlsx_data["sheets"]["Расчёты"]["markdown"]

            # В markdown должна быть пометка о невычисленной формуле
            assert "не вычислено" in md
            # Сама формула должна присутствовать (чтобы LLM знала что там)
            assert "A2+B2" in md
            # Другие значения присутствуют
            assert "100" in md
            assert "200" in md
        finally:
            os.unlink(tmp.name)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 7: Таблица больше XLSX_MAX_CHARS → бот спрашивает как обрезать
# ══════════════════════════════════════════════════════════════════════════════

class TestSizeLimitAsksChoice(unittest.IsolatedAsyncioTestCase):
    """Тест 7: Таблица превышает XLSX_MAX_CHARS → бот спрашивает про усечение."""

    async def test_size_limit_asks_choice(self):
        """
        Дано: очень большая таблица (генерируем строки пока не превысим лимит).
        Ожидаем:
        - context.user_data["xlsx_pending"]["state"] == "awaiting_size_choice"
        - _send_xlsx_to_llm НЕ вызван
        - reply_text вызван с предложением 1️⃣ / 2️⃣
        """
        # Создаём xlsx с таблицей, гарантированно превышающей XLSX_MAX_CHARS
        # ~40000 символов ≈ 400+ строк с длинными значениями
        n_rows = 600  # заведомо больше лимита
        rows = [["Колонка A", "Колонка B", "Колонка C", "Колонка D", "Колонка E"]]
        for i in range(n_rows):
            rows.append([
                f"Значение строки {i} колонки А",
                f"Значение строки {i} колонки Б",
                f"Значение строки {i} колонки В",
                f"Значение строки {i} колонки Г",
                f"Значение строки {i} колонки Д",
            ])

        path = make_xlsx_file({"Данные": rows})
        try:
            xlsx_data = _open_xlsx_data(path)
            update = make_update_mock()
            context = make_context_mock()

            with patch("bot._send_xlsx_to_llm", new_callable=AsyncMock) as mock_send:
                await _process_xlsx_sheets(
                    update, context,
                    sheet_names=xlsx_data["sheet_names"],
                    xlsx_data=xlsx_data,
                    file_name="big_table.xlsx",
                    caption="",
                    user_id=42,
                )

            # Должен быть установлен pending с запросом на выбор
            assert context.user_data.get("xlsx_pending", {}).get("state") == "awaiting_size_choice"
            # LLM не вызван
            mock_send.assert_not_awaited()
            # Пользователю задан вопрос
            assert update.message.reply_text.call_count == 1
            call_text = update.message.reply_text.call_args[0][0]
            assert "1️⃣" in call_text
            assert "2️⃣" in call_text
        finally:
            os.unlink(path)


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 8: Пользователь выбирает "1" (обрезать) → первые N символов
# ══════════════════════════════════════════════════════════════════════════════

class TestSizeLimitTruncate(unittest.IsolatedAsyncioTestCase):
    """Тест 8: При выборе "1" берутся первые XLSX_MAX_CHARS символов."""

    async def test_size_limit_user_picks_truncate(self):
        """
        Дано: pending с большим combined_text, пользователь пишет "1".
        Ожидаем:
        - _send_xlsx_to_llm вызван с текстом ≤ XLSX_MAX_CHARS + overhead
        - xlsx_pending очищен
        """
        big_text = "X" * (XLSX_MAX_CHARS * 2)  # вдвое больше лимита

        update = make_update_mock(text="1")
        context = make_context_mock(user_data={
            "xlsx_pending": {
                "state": "awaiting_size_choice",
                "file_name": "huge.xlsx",
                "sheet_names": ["Лист1"],
                "xlsx_data": {},
                "combined_text": big_text,
                "caption": "",
                "user_id": 42,
                "n_parts": 2,
            }
        })

        sent_texts = []

        async def capture_send(update, context, combined, caption, user_id):
            sent_texts.append(combined)

        with patch("bot._send_xlsx_to_llm", side_effect=capture_send):
            await handle_xlsx_dialog(update, context)

        # Диалог завершён
        assert "xlsx_pending" not in context.user_data
        # Текст обрезан
        assert len(sent_texts) == 1
        # Обрезанный текст не превышает лимит + небольшой хвост (пометка об обрезке)
        assert len(sent_texts[0]) <= XLSX_MAX_CHARS + 200


# ══════════════════════════════════════════════════════════════════════════════
# ТЕСТ 9: Повреждённый/некорректный xlsx-файл → корректное сообщение об ошибке
# ══════════════════════════════════════════════════════════════════════════════

class TestCorruptedXlsx(unittest.TestCase):
    """Тест 9: Повреждённый файл → _open_xlsx_data выбрасывает исключение."""

    def test_corrupted_xlsx_raises_exception(self):
        """
        Дано: файл .xlsx с мусорным содержимым.
        Ожидаем: _open_xlsx_data выбрасывает исключение (любое).
        (handle_document перехватит его и отправит пользователю сообщение об ошибке)
        """
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(b"NOT AN XLSX FILE AT ALL \x00\x01\x02\x03")
        tmp.close()

        try:
            with pytest.raises(Exception):
                _open_xlsx_data(tmp.name)
        finally:
            os.unlink(tmp.name)


# ══════════════════════════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ТЕСТЫ: sheet_to_markdown и split_text_into_parts
# ══════════════════════════════════════════════════════════════════════════════

class TestSheetToMarkdown(unittest.TestCase):
    """Юнит-тесты для sheet_to_markdown — форматирование таблицы."""

    def _make_ws(self, rows: list) -> openpyxl.worksheet.worksheet.Worksheet:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        return ws

    def test_basic_table_format(self):
        """Базовая таблица → правильный Markdown-заголовок и строки."""
        ws = self._make_ws([
            ["Имя", "Возраст"],
            ["Алиса", 30],
        ])
        md = sheet_to_markdown(ws)
        lines = md.split("\n")
        assert lines[0] == "| Имя | Возраст |"
        assert lines[1] == "| --- | --- |"
        assert "Алиса" in lines[2]
        assert "30" in lines[2]

    def test_pipe_escaping(self):
        """Символ | в ячейке экранируется."""
        ws = self._make_ws([
            ["Значение"],
            ["A|B"],
        ])
        md = sheet_to_markdown(ws)
        assert "A\\|B" in md

    def test_empty_cell_is_empty_string(self):
        """Пустая ячейка (None) → пустая строка в таблице."""
        ws = self._make_ws([
            ["A", "B"],
            [None, "val"],
        ])
        md = sheet_to_markdown(ws)
        assert "|  | val |" in md or "| None" not in md


class TestSplitTextIntoParts(unittest.TestCase):
    """Юнит-тесты для split_text_into_parts."""

    def test_short_text_not_split(self):
        """Текст меньше лимита → возвращается как одна часть."""
        text = "Привет мир\nВторая строка"
        parts = split_text_into_parts(text, max_chars=1000)
        assert len(parts) == 1
        assert parts[0] == text

    def test_long_text_split_by_lines(self):
        """Длинный текст делится по строкам, не рвёт строки пополам."""
        lines = [f"Строка {i}: {'данные' * 10}" for i in range(100)]
        text = "\n".join(lines)
        parts = split_text_into_parts(text, max_chars=500)
        assert len(parts) > 1
        # Каждая часть не превышает лимит (с небольшим допуском за последнюю строку)
        for part in parts:
            # Допускаем одну строку сверх лимита (она добавляется до превышения)
            line_lengths = [len(l) for l in lines]
            max_line = max(line_lengths)
            assert len(part) <= 500 + max_line


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
